"""
xlsx_utils.py

Utilidades para integrar insumos al Libro Maestro (Insumo 1: "Inventario de proveedor").

Objetivos:
----------
1. Mantener la estructura del maestro:
   - No perder encabezados.
   - Respetar la "fila plantilla" de fórmulas (normalmente la fila 2).
   - Copiar las fórmulas a las nuevas filas generadas.

2. Integrar insumos de forma controlada:
   - Insumo 2 (Endpoint/Antivirus) → hoja "Antivirus".
   - Insumo 3 (Personal) → hoja "ESTADO_GEN_USUARIO".
   - Insumo 4 (TMP) → hoja "Useraranda_BLOGIK".
   - Insumo 5 (DA) → hoja "Reporte DA".

3. Permitir integración incremental:
   - El insumo de personal **no borra** la hoja completa, sino que:
     - Actualiza filas existentes si la cédula ya está.
     - Inserta nuevas filas si la cédula no está.

4. Hacer backups automáticos para evitar pérdida de datos.

Requisitos:
-----------
- pandas
- openpyxl

Instalación:
------------
pip install pandas openpyxl
"""

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Iterable
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ======================================================
#                UTILIDADES BÁSICAS
# ======================================================

def backup_file(path: str) -> str:
    """
    Crea un backup del archivo Excel antes de sobrescribirlo.

    Ejemplo:
        master.xlsx → master_backup_20250915_103000.xlsx
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(f"No existe el archivo para backup: {path}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = src.with_name(f"{src.stem}_backup_{ts}{src.suffix}")
    shutil.copy2(str(src), str(dst))
    return str(dst)


def _norm_text(s: str) -> str:
    """
    Normaliza un texto:
      - minúsculas
      - sin tildes
      - sin dobles espacios

    Útil para comparar nombres de columnas de manera flexible.
    """
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
    s = re.sub(r"\s+", " ", s)
    return s


def _build_header_map(ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
    """
    Construye un diccionario {nombre_columna_normalizado: número_columna}.
    Permite ubicar en qué columna está cada encabezado en la hoja.
    """
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        key = _norm_text(val)
        if key:
            headers[key] = col
    return headers


def _delete_data_rows(ws: Worksheet, keep_rows: int = 2) -> None:
    """
    Elimina todas las filas de datos, conservando las primeras `keep_rows`.

    Ejemplo:
      keep_rows = 2 → conserva encabezado y plantilla de fórmulas.
    """
    max_r = ws.max_row
    if max_r > keep_rows:
        ws.delete_rows(keep_rows + 1, max_r - keep_rows)


def _copy_formula_row(ws: Worksheet, from_row: int, to_row_start: int, to_row_end: int) -> None:
    """
    Copia fórmulas de una fila plantilla (`from_row`) hacia un rango de filas nuevas.
    """
    if to_row_end < to_row_start:
        return
    for col in range(1, ws.max_column + 1):
        tmpl_val = ws.cell(row=from_row, column=col).value
        if isinstance(tmpl_val, str) and tmpl_val.startswith('='):
            for r in range(to_row_start, to_row_end + 1):
                ws.cell(row=r, column=col).value = tmpl_val


def _set_cell(ws: Worksheet, row: int, col: int, value):
    """Escribe un valor en una celda específica."""
    ws.cell(row=row, column=col, value=value)


def _first_match_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    """
    Busca en el DataFrame la primera columna que coincida con una lista de nombres posibles.
    Devuelve el nombre real de la columna encontrada o None.
    """
    norm_map = { _norm_text(c): c for c in df.columns }
    for cand in candidates:
        k = _norm_text(cand)
        if k in norm_map:
            return norm_map[k]
    return None


def _fecha_from_filename(filename: str) -> Optional[datetime]:
    """
    Extrae una fecha de un nombre de archivo si está embebida:
      - Formato soportado: YYYYMMDD, YYYY-MM-DD, DD-MM-YYYY, etc.
    """
    base = Path(filename).stem
    m = re.search(r'(\d{4})[._-]?(\d{2})[._-]?(\d{2})', base)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d)
        except Exception:
            pass
    m = re.search(r'(\d{2})[._-](\d{2})[._-](\d{4})', base)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return datetime(y, mo, d)
        except Exception:
            pass
    return None


def _clean_cedula(value) -> Optional[str]:
    """
    Limpia un número de cédula:
      - elimina puntos y comas
      - elimina caracteres no numéricos
    """
    if value is None:
        return None
    s = str(value).strip()
    s = s.replace('.', '').replace(',', '')
    s = re.sub(r'\D+', '', s)
    return s or None


def _compose_nombre(df_row: pd.Series, cols: Dict[str, Optional[str]]) -> Optional[str]:
    """
    Construye el nombre completo a partir de las columnas disponibles en un DataFrame.
    Si existe "NOMBRE COMPLETO", lo usa directamente.
    """
    nc = cols.get('nombre_completo')
    if nc and pd.notna(df_row.get(nc)):
        val = str(df_row[nc]).strip()
        if val:
            return val

    # Si no hay columna directa, concatenar nombres y apellidos
    pa = cols.get('primer_apellido')
    sa = cols.get('segundo_apellido')
    pn = cols.get('primer_nombre')
    sn = cols.get('segundo_nombre')
    apellidos = []
    nombres = []
    for key in (pa, sa):
        if key and pd.notna(df_row.get(key)):
            txt = str(df_row[key]).strip()
            if txt:
                apellidos.append(txt)
    for key in (pn, sn):
        if key and pd.notna(df_row.get(key)):
            txt = str(df_row[key]).strip()
            if txt:
                nombres.append(txt)
    if not (apellidos or nombres):
        return None
    return f"{' '.join(apellidos)} {' '.join(nombres)}".strip()


# ======================================================
#           INTEGRACIÓN DE INSUMO 3: PERSONAL
# ======================================================

def integrate_personnel_to_estado(
    master_path: str,
    personnel_path: str,
    keep_rows: int = 2,
    area: Optional[str] = None,
    operacion: Optional[str] = None,
    fecha_archivo: Optional[datetime] = None,
    make_backup: bool = True,
) -> Dict:
    """
    Integra el archivo de personal en la hoja "ESTADO_GEN_USUARIO" del maestro.

    Estrategia incremental:
      - Buscar por CÉDULA.
      - Si ya existe: actualizar nombre, dependencia, área, fecha y estado.
      - Si no existe: agregar una nueva fila al final.
      - No borra toda la hoja.

    Parámetros:
      - master_path: ruta al archivo maestro (Excel).
      - personnel_path: ruta al archivo con datos de personal.
      - keep_rows: cuántas filas conservar (1=encabezado, 2=encabezado+plantilla).
      - area: permite forzar área global (si no se deduce del archivo).
      - operacion: "ingresos", "retiros" o "mixto". Si None, se infiere del nombre del archivo.
      - fecha_archivo: si se pasa, se usa como fecha por defecto.
      - make_backup: si True, hace backup antes de sobrescribir.

    Retorna:
      Diccionario con resumen:
        {added, updated, skipped, backup, last_row, ...}
    """
    sheet_name = "ESTADO_GEN_USUARIO"

    # Validaciones iniciales
    if not os.path.exists(master_path):
        raise FileNotFoundError(master_path)
    if not os.path.exists(personnel_path):
        raise FileNotFoundError(personnel_path)

    # Backup opcional
    backup = backup_file(master_path) if make_backup else None

    # Cargar maestro
    wb = load_workbook(master_path, data_only=False, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el maestro.")
    ws = wb[sheet_name]

    headers_map = _build_header_map(ws, header_row=1)

    # Ubicar columnas destino
    dst_cedula = headers_map.get(_norm_text("CEDULA"))
    dst_nombre = headers_map.get(_norm_text("NOMBRE"))
    dst_dependencia = headers_map.get(_norm_text("DEPENDENCIA"))
    dst_area = headers_map.get(_norm_text("AREA"))
    dst_estado = headers_map.get(_norm_text("ESTADO"))
    dst_ing_ret = (
        headers_map.get(_norm_text("INGRESO/RETIRO"))
        or headers_map.get(_norm_text("INGRESO"))
        or headers_map.get(_norm_text("FECHA"))
    )

    if not dst_cedula:
        raise ValueError("No se encontró la columna 'CEDULA' en la hoja destino. Es obligatoria.")

    # Leer archivo de personal
    df = pd.read_excel(personnel_path, engine='openpyxl')

    # Buscar columnas fuente
    col_doc = _first_match_column(df, ["Documento", "Cédula", "Cedula", "NUMERO DOCUMENTO", "No. documento"])
    col_nombre_completo = _first_match_column(df, ["NOMBRE COMPLETO", "Nombre completo", "Nombres y apellidos"])
    col_primer_nombre = _first_match_column(df, ["Primer nombre", "Nombre 1"])
    col_segundo_nombre = _first_match_column(df, ["Segundo nombre", "Nombre 2"])
    col_primer_apellido = _first_match_column(df, ["Primer apellido", "Apellido 1"])
    col_segundo_apellido = _first_match_column(df, ["Segundo apellido", "Apellido 2"])
    col_dependencia = _first_match_column(df, ["DEPENDENCIA", "Centro de costos"])
    col_area = _first_match_column(df, ["AREA", "REGIONAL"])
    col_fec_term = _first_match_column(df, ["FECHA TERMINACIÓN", "FECHA TERMINACION", "fecha fin"])
    col_fec_ini = _first_match_column(df, ["FECHA INICIO", "fecha ingreso"])

    # Inferir parámetros globales si no se pasan
    filename = Path(personnel_path).name
    fname_norm = _norm_text(filename)

    def _guess_operacion() -> str:
        if operacion:
            return operacion.lower()
        if any(x in fname_norm for x in ["retiro", "terminacion", "fin"]):
            return "retiros"
        if "ingres" in fname_norm:
            return "ingresos"
        return "mixto"

    def _guess_area() -> str:
        if area:
            return area
        if "fomag" in fname_norm:
            return "FOMAG"
        if "m.c" in fname_norm or "mc" in fname_norm:
            return "M.C"
        if "apre" in fname_norm or "pract" in fname_norm:
            return "APRE Y PRACT"
        if "mision" in fname_norm:
            return "FIDU MISIÓN"
        return "FIDU PLANTA"

    def _estado_from(op: str, ar: str, row_has_term: bool, row_has_ini: bool) -> str:
        """Deducción del ESTADO basado en operación, área y fechas."""
        op = (op or "").lower()
        ar_up = (ar or "").upper()
        if "reti" in op or "termin" in op:
            base = "RETIRADO"
        elif "ingre" in op:
            base = "ACTIVO"
        else:
            base = "RETIRADO" if row_has_term else "ACTIVO"
        suf = "FIDU PLANTA"
        if "FOMAG" in ar_up:
            suf = "FOMAG"
        elif "M.C" in ar_up:
            suf = "M.C"
        elif "APRE" in ar_up or "PRACT" in ar_up:
            suf = "APRE Y PRACT"
        elif "MISION" in _norm_text(ar_up):
            suf = "FIDU MISIÓN"
        return f"{base} {suf}"

    op_global = _guess_operacion()
    area_global = _guess_area()
    if fecha_archivo is None:
        fecha_archivo = _fecha_from_filename(filename)

    # Construir índice de cédulas existentes
    start_row = keep_rows + 1
    existing_map: Dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=dst_cedula).value
        c = _clean_cedula(cell_val)
        if c and c not in existing_map:
            existing_map[c] = r

    # Contadores
    added, updated, skipped = 0, 0, 0
    appended_rows: List[int] = []

    nombre_cols = {
        "nombre_completo": col_nombre_completo,
        "primer_apellido": col_primer_apellido,
        "segundo_apellido": col_segundo_apellido,
        "primer_nombre": col_primer_nombre,
        "segundo_nombre": col_segundo_nombre,
    }

    # Procesar filas del insumo
    for _, row in df.iterrows():
        # Cédula
        ced = _clean_cedula(row.get(col_doc)) if col_doc else None
        if not ced:
            skipped += 1
            continue

        # Valores
        nombre_val = _compose_nombre(row, nombre_cols)
        dep_val = str(row.get(col_dependencia)).strip() if col_dependencia and pd.notna(row.get(col_dependencia)) else None
        area_val = str(row.get(col_area)).strip() if col_area and pd.notna(row.get(col_area)) else area_global
        fecha_val = row.get(col_fec_term) if col_fec_term and pd.notna(row.get(col_fec_term)) else None
        if not fecha_val and fecha_archivo:
            fecha_val = fecha_archivo.date()

        # Estado calculado
        row_has_term = col_fec_term and pd.notna(row.get(col_fec_term))
        row_has_ini = col_fec_ini and pd.notna(row.get(col_fec_ini))
        estado_val = _estado_from(op_global, area_val, row_has_term, row_has_ini)

        if ced in existing_map:
            # Actualizar fila existente
            r = existing_map[ced]
            if dst_nombre and nombre_val:
                _set_cell(ws, r, dst_nombre, nombre_val)
            if dst_dependencia and dep_val:
                _set_cell(ws, r, dst_dependencia, dep_val)
            if dst_area and area_val:
                _set_cell(ws, r, dst_area, area_val)
            if dst_ing_ret and fecha_val:
                _set_cell(ws, r, dst_ing_ret, fecha_val)
            if dst_estado:
                _set_cell(ws, r, dst_estado, estado_val)
            updated += 1
        else:
            # Insertar nueva fila
            new_row = max(ws.max_row + 1, start_row)
            _set_cell(ws, new_row, dst_cedula, ced)
            if dst_nombre and nombre_val:
                _set_cell(ws, new_row, dst_nombre, nombre_val)
            if dst_dependencia and dep_val:
                _set_cell(ws, new_row, dst_dependencia, dep_val)
            if dst_area and area_val:
                _set_cell(ws, new_row, dst_area, area_val)
            if dst_ing_ret and fecha_val:
                _set_cell(ws, new_row, dst_ing_ret, fecha_val)
            if dst_estado:
                _set_cell(ws, new_row, dst_estado, estado_val)
            appended_rows.append(new_row)
            existing_map[ced] = new_row
            added += 1

    # Copiar fórmulas desde la fila plantilla a las filas nuevas
    if keep_rows >= 2 and appended_rows:
        _copy_formula_row(ws, from_row=2, to_row_start=min(appended_rows), to_row_end=max(appended_rows))

    # Guardar maestro actualizado
    wb.save(master_path)

    return {
        "sheet": sheet_name,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "backup": backup,
        "last_row": ws.max_row,
        "keep_rows": keep_rows,
    }


# ======================================================
#       INTEGRACIÓN DIRECTA (TMP, DA, ETC.)
# ======================================================

def replace_sheet_with_df(master_path: str, sheet_name: str, df: pd.DataFrame, keep_rows: int = 1) -> Dict:
    """
    Reemplaza los datos de una hoja entera por el contenido de un DataFrame.

    - Respeta las filas iniciales (keep_rows).
    - Copia las fórmulas de la fila 2 a las nuevas filas.
    """
    if not os.path.exists(master_path):
        raise FileNotFoundError(master_path)

    backup = backup_file(master_path)

    wb = load_workbook(master_path, data_only=False, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el maestro.")
    ws = wb[sheet_name]

    # Borrar datos previos pero conservar encabezados y plantilla
    _delete_data_rows(ws, keep_rows=keep_rows)

    # Insertar nuevos datos desde DataFrame
    start_row = keep_rows + 1
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Copiar fórmulas desde la fila plantilla si corresponde
    if keep_rows >= 2 and df.shape[0] > 0:
        _copy_formula_row(ws, from_row=2, to_row_start=start_row, to_row_end=start_row + df.shape[0] - 1)

    wb.save(master_path)

    return {
        "sheet": sheet_name,
        "rows_written": df.shape[0],
        "backup": backup,
        "last_row": ws.max_row,
        "keep_rows": keep_rows,
    }